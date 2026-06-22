import json
import os
import tempfile
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import openpyxl

from excel_bridge import parse_flow, parse_tabular


BASE_DIR = Path(__file__).resolve().parent
HOST = os.environ.get("HOST", "0.0.0.0")
PORT = int(os.environ.get("PORT", "8000"))
TEMPLATE_CANDIDATES = [
    Path(path)
    for path in [
        os.environ.get("PNL_TEMPLATE_PATH", ""),
        r"C:\Users\admin\Documents\Jarvis\uploads\损益测算\科目分类和预估方法-V6.xlsx",
        BASE_DIR / "template-layout.xlsx",
    ]
    if path
]


def _json_bytes(payload):
    return json.dumps(payload, ensure_ascii=False).encode("utf-8")


def _normalize_text(value):
    return "" if value is None else str(value).strip()


def _normalize_name(value):
    return "" if value is None else str(value)


def _normalize_number(value):
    try:
        return float(value)
    except Exception:
        return 0.0


def _normalize_date(value):
    return "" if not value else str(value)[:10]


def _normalize_month(value):
    return "" if not value else str(value)[:7]


def _normalize_tabular_rows(payload):
    rows = payload.get("rows", [])
    normalized = []
    for row in rows:
        normalized_row = {
            "dt": _normalize_date(row.get("dt")),
            "fcst_dept_code": _normalize_text(row.get("fcst_dept_code")),
            "fcst_dept_name": _normalize_text(row.get("fcst_dept_name")),
            "fcst_acc_name": _normalize_text(row.get("fcst_acc_name")),
            "amount": _normalize_number(row.get("amount")),
            "fee_type": _normalize_text(row.get("fee_type")),
            "num": _normalize_number(row.get("num")),
            "display_name": _normalize_text(row.get("display_name")),
            "fcst_acc_code": _normalize_text(row.get("fcst_acc_code")),
            "org_category": _normalize_text(row.get("org_category")),
            "dept_label": _normalize_text(row.get("dept_label")),
            "group_label": _normalize_text(row.get("group_label")),
            "update_dt": _normalize_text(row.get("update_dt")),
            "region": _normalize_text(row.get("region")),
            "sheet": _normalize_text(row.get("sheet")),
        }
        if not (
            normalized_row["dt"]
            or normalized_row["fcst_acc_code"]
            or normalized_row["fcst_acc_name"]
            or normalized_row["display_name"]
        ):
            continue
        normalized.append(normalized_row)
    return normalized


def _build_dataset_summary(rows, scope):
    months = sorted({r["dt"] for r in rows if r["dt"]})
    max_dt = months[-1] if months else ""
    display_name = next((r["display_name"] for r in rows if r["display_name"]), "")
    sheets = sorted({r["sheet"] for r in rows if r.get("sheet")})
    return {
        "scope": scope,
        "rowCount": len(rows),
        "displayName": display_name,
        "maxDt": max_dt,
        "sheetCount": len(sheets),
        "sheets": sheets,
        "rows": rows,
    }


def _parse_workbook_bytes(body, scope):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(body)
        temp_path = Path(tmp.name)

    try:
        workbook = openpyxl.load_workbook(temp_path, data_only=True, read_only=True)
        if scope == "flow":
            sheet = workbook[workbook.sheetnames[0]]
            payload = parse_flow(sheet)
            parsed_rows = payload.get("rows", [])
            rows = [
                {
                    "month": _normalize_month(row.get("month") or row.get("dt")),
                    "revenue": _normalize_number(row.get("revenue")),
                    "channels": {
                        str(key): _normalize_number(value)
                        for key, value in (row.get("channels") or {}).items()
                    },
                }
                for row in parsed_rows
            ]
            return {
                "rowCount": len(rows),
                "rows": rows,
                "matchedChannels": payload.get("matchedChannels", []),
                "unmatchedChannels": payload.get("unmatchedChannels", []),
                "sourceFormat": payload.get("sourceFormat", ""),
                "sheet": payload.get("sheet", ""),
            }

        combined_rows = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            payload = parse_tabular(sheet)
            parsed_rows = payload.get("rows", [])
            if not parsed_rows:
                continue
            combined_rows.extend([
                {
                    **row,
                    "sheet": payload.get("sheet") or sheet_name,
                }
                for row in parsed_rows
            ])

        normalized = _normalize_tabular_rows({"rows": combined_rows})
        return _build_dataset_summary(normalized, scope)
    finally:
        try:
            temp_path.unlink(missing_ok=True)
        except Exception:
            pass


def _resolve_template_path():
    for candidate in TEMPLATE_CANDIDATES:
        if candidate.exists():
            return candidate
    return None


def _load_template_layout():
    template_path = _resolve_template_path()
    if not template_path:
        return {"rows": []}

    workbook = openpyxl.load_workbook(template_path, data_only=True, read_only=False)
    sheet = workbook[workbook.sheetnames[0]]
    rows = []

    for row_idx in range(3, sheet.max_row + 1):
        code = _normalize_text(sheet.cell(row=row_idx, column=2).value)
        en_name = _normalize_name(sheet.cell(row=row_idx, column=3).value)
        name_cell = sheet.cell(row=row_idx, column=4)
        name = _normalize_name(name_cell.value) or en_name
        try:
            indent = int(name_cell.alignment.indent or 0)
        except Exception:
            indent = 0
        if not code and not name:
            continue
        if code == "fcst_acc_code":
            continue
        rows.append({"code": code, "name": name, "enName": en_name, "indent": indent})

    return {"rows": rows}


class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(BASE_DIR), **kwargs)

    def _apply_cors_headers(self):
        origin = self.headers.get("Origin") or "*"
        self.send_header("Access-Control-Allow-Origin", origin)
        self.send_header("Vary", "Origin")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Access-Control-Allow-Methods", "GET,POST,OPTIONS")

    def _send_json(self, status, payload):
        body = _json_bytes(payload)
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self._apply_cors_headers()
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        self.send_response(204)
        self._apply_cors_headers()
        self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/health":
            self._send_json(200, {"ok": True})
            return
        if parsed.path == "/api/template-layout":
            self._send_json(200, _load_template_layout())
            return
        if parsed.path == "/":
            self.path = "/index.html"
        return super().do_GET()

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path != "/api/upload-data":
            self._send_json(404, {"error": "not_found"})
            return

        query = parse_qs(parsed.query)
        scope = (query.get("scope") or [""])[0]
        if scope not in {"post", "pre", "allocation", "flow"}:
            self._send_json(400, {"error": "unknown_scope"})
            return

        content_length = int(self.headers.get("Content-Length", "0"))
        if content_length <= 0:
            self._send_json(400, {"error": "empty_body"})
            return

        body = self.rfile.read(content_length)
        try:
            payload = _parse_workbook_bytes(body, scope)
            self._send_json(200, payload)
        except Exception as exc:
            self._send_json(500, {"error": "parse_failed", "message": str(exc)})

    def end_headers(self):
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def log_message(self, *_args):
        return


def main():
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    print(f"http://{HOST}:{PORT}")
    server.serve_forever()


if __name__ == "__main__":
    main()
