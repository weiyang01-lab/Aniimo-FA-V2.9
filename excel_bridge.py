import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime

import openpyxl
from openpyxl.utils.datetime import from_excel


def normalize_cell(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return None
    return value


def normalize_date_like(value):
    if isinstance(value, (datetime, date)):
        return normalize_cell(value)
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date().isoformat()
        except Exception:
            return str(value)
    if value is None:
        return ""
    return str(value)[:10]


FLOW_CHANNEL_ORDER = ["globalMb", "globalPc", "globalConsole", "cnMb", "cnPc"]
FLOW_CHANNEL_LABELS = {
    "globalMb": "Global_mb",
    "globalPc": "Global_pc",
    "globalConsole": "Global_主机",
    "cnMb": "CN_mb",
    "cnPc": "CN_pc",
}


def normalize_match_token(value):
    return re.sub(r"[\s_\-()（）/\\:：]+", "", str(value or "").strip().lower())


def looks_like_date_value(value):
    text = normalize_date_like(value)
    if not text:
        return False
    return bool(
        re.match(r"^\d{4}-\d{2}-\d{2}$", text)
        or re.match(r"^\d{4}/\d{1,2}/\d{1,2}$", str(value or "").strip())
        or re.match(r"^\d{4}-\d{2}$", text)
    )


def normalize_flow_month(value):
    text = normalize_date_like(value)
    if not text:
        return ""
    matched = re.match(r"^(\d{4})[-/](\d{1,2})", text)
    if matched:
        return f"{matched.group(1)}-{int(matched.group(2)):02d}"
    return text[:7]


def to_number(value):
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def recognize_flow_channel(*values):
    raw = " ".join(str(value or "").strip() for value in values if str(value or "").strip())
    token = normalize_match_token(raw)
    if not token:
        return None

    exact_aliases = {
        "globalmb": "globalMb",
        "globalmobile": "globalMb",
        "global_pc": "globalPc",
        "globalpc": "globalPc",
        "globalconsole": "globalConsole",
        "global主机": "globalConsole",
        "cnmb": "cnMb",
        "cnmobile": "cnMb",
        "cnpc": "cnPc",
    }
    if token in exact_aliases:
        return exact_aliases[token]

    has_global = any(marker in token for marker in ["global", "海外", "intl", "international", "worldwide"])
    has_cn = any(marker in token for marker in ["cn", "china", "中国", "国内", "国服", "大陆"])
    has_console = any(marker in token for marker in ["console", "主机", "playstation", "ps", "xbox", "switch"])
    has_pc = any(marker in token for marker in ["pc", "steam", "epic", "wegame", "客户端", "端游"])
    has_mobile = any(
        marker in token
        for marker in [
            "mb",
            "mobile",
            "ios",
            "android",
            "googleplay",
            "appstore",
            "apk",
            "taptap",
            "huawei",
            "oppo",
            "vivo",
            "xiaomi",
            "bilibili",
            "应用商店",
            "手游",
        ]
    )

    if has_console and has_cn:
        return "globalConsole"
    if has_console:
        return "globalConsole"
    if has_pc and has_cn:
        return "cnPc"
    if has_pc and has_global:
        return "globalPc"
    if has_mobile and has_cn:
        return "cnMb"
    if has_mobile and has_global:
        return "globalMb"

    if "wegame" in token:
        return "cnPc"
    if any(marker in token for marker in ["tap", "huawei", "oppo", "vivo", "xiaomi", "bilibili"]):
        return "cnMb"
    if any(marker in token for marker in ["steam", "epic"]):
        return "globalPc"
    if any(marker in token for marker in ["playstation", "xbox", "switch", "主机"]):
        return "globalConsole"
    return None


def empty_flow_channel_map():
    return {key: 0.0 for key in FLOW_CHANNEL_ORDER}


def aggregate_flow_monthly(entries, source_format, unmatched_channels=None):
    monthly = defaultdict(lambda: {"revenue": 0.0, "channels": empty_flow_channel_map()})
    matched_channels = set()

    for entry in entries:
        month = normalize_flow_month(entry.get("dt") or entry.get("month"))
        channel = entry.get("channel")
        revenue = to_number(entry.get("revenue"))
        if not month or not channel:
            continue
        matched_channels.add(channel)
        monthly[month]["channels"][channel] += revenue
        monthly[month]["revenue"] += revenue

    rows = [
        {
            "month": month,
            "revenue": round(payload["revenue"], 4),
            "channels": {key: round(value, 4) for key, value in payload["channels"].items()},
        }
        for month, payload in sorted(monthly.items())
    ]

    return {
        "sheet": "",
        "sourceFormat": source_format,
        "matchedChannels": [FLOW_CHANNEL_LABELS[key] for key in FLOW_CHANNEL_ORDER if key in matched_channels],
        "unmatchedChannels": sorted(set(unmatched_channels or [])),
        "rows": rows,
    }


def find_date_column(headers, data_rows):
    header_tokens = [normalize_match_token(header) for header in headers]
    for idx, token in enumerate(header_tokens):
        if any(keyword in token for keyword in ["dt", "date", "day", "日期", "时间", "统计日期", "流水日期"]):
            return idx

    best_idx = None
    best_score = 0
    for idx in range(len(headers)):
        score = sum(1 for row in data_rows[:20] if idx < len(row) and looks_like_date_value(row[idx]))
        if score > best_score:
            best_idx = idx
            best_score = score
    return best_idx if best_score >= 2 else None


def parse_flow_wide(headers, data_rows, date_col, recognized_channels):
    entries = []
    for row in data_rows:
        if date_col >= len(row):
            continue
        dt = row[date_col]
        if not looks_like_date_value(dt):
            continue
        for idx, channel in recognized_channels.items():
            if idx >= len(row):
                continue
            revenue = to_number(row[idx])
            if revenue == 0:
                continue
            entries.append({"dt": dt, "channel": channel, "revenue": revenue})

    payload = aggregate_flow_monthly(entries, "wide_columns")
    payload["sheet"] = ""
    return payload


def parse_flow_long(headers, data_rows):
    normalized_headers = [normalize_match_token(header) for header in headers]
    date_col = find_date_column(headers, data_rows)
    channel_col = None
    region_col = None
    platform_col = None
    amount_col = None

    for idx, token in enumerate(normalized_headers):
        if channel_col is None and any(keyword in token for keyword in ["channel", "渠道", "来源", "platform", "平台"]):
            channel_col = idx
            continue
        if region_col is None and any(keyword in token for keyword in ["region", "地区", "区域", "市场", "国家", "locale"]):
            region_col = idx
            continue
        if platform_col is None and any(keyword in token for keyword in ["terminal", "device", "端", "os", "系统", "商店", "store"]):
            platform_col = idx
            continue
        if amount_col is None and any(keyword in token for keyword in ["revenue", "amount", "流水", "收入", "金额", "gmv", "bookings"]):
            amount_col = idx

    if amount_col is None:
        numeric_scores = []
        for idx in range(len(headers)):
            if idx in {date_col, channel_col, region_col, platform_col}:
                continue
            score = sum(1 for row in data_rows[:20] if idx < len(row) and isinstance(row[idx], (int, float)))
            numeric_scores.append((score, idx))
        numeric_scores.sort(reverse=True)
        amount_col = numeric_scores[0][1] if numeric_scores and numeric_scores[0][0] > 0 else None

    entries = []
    unmatched_channels = set()
    for row in data_rows:
        if date_col is None or amount_col is None:
            continue
        if date_col >= len(row) or amount_col >= len(row):
            continue
        dt = row[date_col]
        if not looks_like_date_value(dt):
            continue

        channel_parts = []
        if channel_col is not None and channel_col < len(row):
            channel_parts.append(row[channel_col])
        if region_col is not None and region_col < len(row):
            channel_parts.append(row[region_col])
        if platform_col is not None and platform_col < len(row):
            channel_parts.append(row[platform_col])

        channel = recognize_flow_channel(*channel_parts)
        revenue = to_number(row[amount_col])
        if not channel:
            raw_label = " / ".join(str(part or "").strip() for part in channel_parts if str(part or "").strip())
            if raw_label:
                unmatched_channels.add(raw_label)
            continue
        if revenue == 0:
            continue
        entries.append({"dt": dt, "channel": channel, "revenue": revenue})

    payload = aggregate_flow_monthly(entries, "daily_rows", unmatched_channels=unmatched_channels)
    payload["sheet"] = ""
    return payload


def parse_tabular(sheet):
    rows = [[normalize_cell(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    rows = [row for row in rows if any(cell not in (None, "") for cell in row)]
    if not rows:
        return {"sheet": sheet.title, "headers": [], "rows": []}

    if looks_like_display_matrix(rows):
        return parse_display_matrix(sheet, rows)

    headers = [str(cell or "").strip() for cell in rows[0]]
    data_rows = []
    for row in rows[1:]:
        record = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[idx] if idx < len(row) else None
        data_rows.append(record)
    return {"sheet": sheet.title, "headers": headers, "rows": data_rows}


def normalize_month_header(value):
    if isinstance(value, (datetime, date)):
        return normalize_cell(value)
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date().isoformat()
        except Exception:
            return ""
    text = str(value or "").strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}$", text):
        return text
    if re.match(r"^\d{4}/\d{1,2}/\d{1,2}$", text):
        year, month, day = text.split("/")
        return f"{year}-{int(month):02d}-{int(day):02d}"
    return ""


def looks_like_display_matrix(rows):
    if len(rows) < 4:
        return False
    if len(rows[2]) < 3 or len(rows[3]) < 2:
        return False

    first_code = str(rows[3][0] or "").strip()
    first_month = normalize_month_header(rows[2][2] if len(rows[2]) > 2 else None)
    return bool(re.match(r"^[A-Za-z]\d", first_code) and first_month)


def parse_display_matrix(sheet, rows):
    display_name = ""
    if len(rows[0]) > 1 and rows[0][1] is not None:
        display_name = str(rows[0][1]).strip()

    month_columns = []
    for idx, value in enumerate(rows[2][2:], start=2):
        month = normalize_month_header(value)
        if month:
            month_columns.append((idx, month))

    records = []
    for row in rows[3:]:
        code = str(row[0] or "").strip() if len(row) > 0 else ""
        name = str(row[1] or "") if len(row) > 1 else ""
        if not code and not name:
            continue

        for idx, month in month_columns:
            value = row[idx] if idx < len(row) else None
            if value in (None, ""):
                continue
            try:
                amount = float(value)
            except Exception:
                continue
            records.append(
                {
                    "dt": month,
                    "display_name": display_name,
                    "fcst_acc_code": code,
                    "fcst_acc_name": name,
                    "amount": amount,
                }
            )

    return {
        "sheet": sheet.title,
        "headers": ["dt", "display_name", "fcst_acc_code", "fcst_acc_name", "amount"],
        "rows": records,
    }


def parse_flow(sheet):
    rows = [[normalize_cell(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    rows = [row for row in rows if any(cell not in (None, "") for cell in row)]
    if not rows:
        return {"sheet": sheet.title, "rows": [], "matchedChannels": [], "unmatchedChannels": [], "sourceFormat": "empty"}

    headers = [str(cell or "").strip() for cell in rows[0]]
    data_rows = rows[1:]
    date_col = find_date_column(headers, data_rows)
    recognized_channels = {
        idx: channel
        for idx, header in enumerate(headers)
        if idx != date_col and (channel := recognize_flow_channel(header))
    }

    if date_col is not None and len(recognized_channels) >= 2:
        payload = parse_flow_wide(headers, data_rows, date_col, recognized_channels)
        payload["sheet"] = sheet.title
        return payload

    payload = parse_flow_long(headers, data_rows)
    payload["sheet"] = sheet.title
    return payload


def main():
    if len(sys.argv) < 3:
        raise SystemExit("Usage: excel_bridge.py <mode> <path>")

    mode = sys.argv[1]
    workbook_path = sys.argv[2]
    workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    if mode == "flow":
        payload = parse_flow(sheet)
    else:
        payload = parse_tabular(sheet)

    print(json.dumps(payload, ensure_ascii=False))


if __name__ == "__main__":
    main()
